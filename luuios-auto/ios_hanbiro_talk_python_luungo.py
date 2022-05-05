#import time, sys, unittest, random, json, requests, openpyxl, testlink
import time, json, random, platform,openpyxl
from datetime import datetime
from appium import webdriver
from random import randint
from selenium.webdriver.common.desired_capabilities import DesiredCapabilities
from selenium.webdriver.common.by import By
from selenium.webdriver import ActionChains
from selenium.webdriver.common.action_chains import ActionChains
from selenium.common.exceptions import TimeoutException, WebDriverException, NoSuchElementException
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait, Select
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.alert import Alert
import time
import inspect
from appium.webdriver.common.touch_action import TouchAction
from sys import exit
from openpyxl import Workbook


class objects:
    now = datetime.now()
    year = now.strftime("%Y")
    month = now.strftime("%m")
    day = now.strftime("%d")
    time1 = now.strftime("%H:%M:%S")
    date_time = now.strftime("%Y/%m/%d, %H:%M:%S")
    date_id = date_time.replace("/", "").replace(", ", "").replace(":", "")[2:]
    testcase_pass = "Test case status: pass"
    testcase_fail = "Test case status: fail"


#now = datetime.now()
#date_time = now.strftime("%Y/%m/%d, %H:%M:%S")
#date_id = date_time.replace("/", "").replace(", ", "").replace(":", "")[2:]

if platform == "linux" or platform == "linux2":
    local = "/home/oem/groupware-auto-test"
    json_file = local + "/appium_hanbiro_talk_android.json"
    with open(json_file) as json_data_file:
        data = json.load(json_data_file)
    log_folder = "/Log/"
    execution_log = local + log_folder + "hanbiro_talk_android_execution_log_" + str(objects.date_id) + ".txt"
    fail_log = execution_log.replace("hanbiro_talk_android_execution_log_", "fail_log_")
    error_log = execution_log.replace("hanbiro_talk_android_execution_log_", "error_log_") 
else :
    local = "/Users/hanbiro/Desktop/luuios"
    json_file = local + "/appium_hanbiro_talk_ios.json"
    with open(json_file) as json_data_file:
        data = json.load(json_data_file)
    log_folder = "/Log/"
    execution_log = local + log_folder + "hanbiro_talk_android_execution_log_" + str(objects.date_id) + ".txt"
    fail_log = execution_log.replace("hanbiro_talk_android_execution_log_", "fail_log_")
    error_log = execution_log.replace("hanbiro_talk_android_execution_log_", "error_log_")


testcase_log = local + log_folder + "testcase_hanbiro_talk_android_" + str(objects.date_id) + ".xlsx"   

logs = [execution_log,fail_log,error_log,testcase_log]
for log in logs: 
    if".txt" in log:
        open(log,"x").close()
    else:
        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1).value = "Menu"
        ws.cell(row=1, column=2).value = "Sub-Menu"
        ws.cell(row=1, column=3).value = "Test Case Name"
        ws.cell(row=1, column=4).value = "Status"
        ws.cell(row=1, column=5).value = "Description"
        ws.cell(row=1, column=6).value = "Date"
        ws.cell(row=1, column=7).value = "Tester"
        wb.save(log)

# Connect to Appium with the below desire capabilities
# http://appium.io/docs/en/writing-running-appium/caps/

APPIUM_PORT = '4723'
udid = 'bc86e429485c13f34837866fde36e7ed55646317'
app_path = 'Users/hanbiro/Desktop/luuios/HanbiroTalk.ipa'
command_executor ='http://127.0.0.1:%s/wd/hub' % APPIUM_PORT

desired_capabilities = {
    'orientation' :'LANDSCAPE',
    "deviceName": "Hanbiro Iphone",
    "platformVersion": "12.5.5",
    "platformName": "IOS",
    "udid": udid,
    "app": app_path,
    "sendKeyStrategy":"setValue"
}
driver = webdriver.Remote(command_executor,desired_capabilities)





now = datetime.now()
mail_title = "Mail App is write at"  + str(now)

# Input information for log-in

#with open("D:\\File_Du_Lieu\\Selenium\\Selenium_python\\Selenium_python\\Appium-pyhton\\LuuNgo_Appium\\appium_mail_app.json") as json_data_file:
#    data = json.load(json_data_file)

def Logging(msg):
    print(msg)
    log_msg = open(execution_log, "a")
    log_msg.write(str(msg) + "\n")
    log_msg.close()

def ValidateFailResultAndSystem(fail_msg):
    Logging(fail_msg)
    append_fail_result = open(fail_log, "a")
    append_fail_result.write("[FAILED TEST CASE] " + str(fail_msg) + "\n")
    append_fail_result.close()
    
time.sleep(1)



def TestCase_LogResult(menu, sub_menu, testcase, status, description, tester):
    Logging("description")
    if status=="Pass":
        print(objects.testcase_pass)
    else:
        print(objects.testcase_fail)

    wb = openpyxl.load_workbook(testcase_log)
    current_sheet = wb.active
    start_row = len(list(current_sheet.rows))+1
    current_sheet.cell(row=start_row,column=1).value=menu
    current_sheet.cell(row=start_row,column=2).value=sub_menu
    current_sheet.cell(row=start_row,column=3).value=testcase
    current_sheet.cell(row=start_row,column=4).value=status
    current_sheet.cell(row=start_row,column=5).value=description
    current_sheet.cell(row=start_row,column=6).value=objects.date_time
    current_sheet.cell(row=start_row,column=7).value= tester
    
    wb.save(testcase_log)




def log_in_hanbiro_talk_ios():
    time.sleep(3)
    Logging("----------------1. Log In Hanbiro Talk-----------------------------")
    #check_crash= WebDriverWait(driver, 100).until(EC.presence_of_element_located((By.XPATH, data["domain_input"])))
    #check_crash= WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["domain_input"])))
    try:
        check_crash = driver.find_element_by_ios_class_chain(data["domain_input"])
        if  check_crash.is_displayed():
            Logging("=>=> No Crash App") 
        else:
            Logging("=>=> Crash App")
            exit(0)
    except WebDriverException:
        ValidateFailResultAndSystem("=>=> Crash App")
        exit(0)
    time.sleep(1)
    try:
        txt_domain= driver.find_element_by_ios_class_chain(data["domain_input"])
        if  txt_domain.is_displayed():
            Logging("=>=> No Crash App")
            txt_domain.send_keys(data["login_page"])
            Logging("1.Input Domain") 
        else:
            ValidateFailResultAndSystem("=>=> Crash App")
            exit(0)
    except WebDriverException:
        ValidateFailResultAndSystem("=>=> Crash App")
        exit(0)
    time.sleep(1)
    
    try:
        #username = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["txt_id"])))
        username = driver.find_element_by_ios_class_chain(data["txt_id"])
        if  username.is_displayed():
            Logging("=>=> No Crash App") 
            username.send_keys(data["input_user_user"])
            Logging("2.Input ID")
        else:
            Logging("=>=> Crash App")
            exit(0)
    except WebDriverException:
        ValidateFailResultAndSystem("=>=> Crash App")
        exit(0)
    time.sleep(1)
    try:
        #password = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["txt_pw"])))
        password = driver.find_element_by_ios_class_chain(data["txt_pw"])
        if  password.is_displayed():
            Logging("=>=> No Crash App") 
            password.send_keys(data["input_pw"])
            Logging("3.Input Password")
        else:
            Logging("=> Crash App")
            exit(0)
    except WebDriverException:
        ValidateFailResultAndSystem("=>=> Crash App")
        exit(0)
    
    button_log_in = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["btn_log_in"])))
    button_log_in.click()
    Logging("4.Click Log In button")
    time.sleep(7)
    
    try:
        check_crash_app_log_in_hanbiro_talk = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["hanbiro_talk_ios"]["check_crash_app_log_in"])))
        if  check_crash_app_log_in_hanbiro_talk.is_displayed():
            Logging("=> No Crash App") 
            TestCase_LogResult(**data["testcase_result"]["hanbiro_talk_ios"]["log_in_hanbiro_talk"]["pass"])
        else:
            ValidateFailResultAndSystem("=> Crash App")
            TestCase_LogResult(**data["testcase_result"]["hanbiro_talk_ios"]["log_in_hanbiro_talk"]["fail"])
            exit(0)
    except WebDriverException:
        ValidateFailResultAndSystem("=> Crash App")
        exit(0)
    
def hanbiro_talk_ios_organization_search():
    Logging("------------------------------------------------------Search Organization------------------------------------------------------")
    time.sleep(1)
    try:
        check_crash_app_log_in_hanbiro_talk = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["hanbiro_talk_ios"]["check_crash_app_log_in"])))
        check_crash_app_log_in_hanbiro_talk.click()
        time.sleep(1)
        Logging("1. Click Company successfully")
        time.sleep(1)
        txt_search_contact_org = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["hanbiro_talk_ios"]["txt_search_contact"])))
        if  txt_search_contact_org.is_displayed():
            Logging("=> No Crash App")        
            txt_search_contact_org123= WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["hanbiro_talk_ios"]["txt_search_contact"])))
            txt_search_contact_org123.click()
            time.sleep(2)
            input_search_contact_org=driver.find_element_by_ios_class_chain(data["hanbiro_talk_ios"]["txt_input_search_contact"])
            input_search_contact_org.click()
            input_search_contact_org.send_keys(data["hanbiro_talk_ios"]["search_contact_org_talk"]+"\n")
            Logging("Input user => pass")
            time.sleep(1)
        else:
                ValidateFailResultAndSystem("=> Crash App")
                exit(0)
    except WebDriverException:
        ValidateFailResultAndSystem("=> Crash App")
        exit(0)
    time.sleep(3)
    Logging("-----------------------Select Contact------------------------")   
    try:
        select_contact_org=driver.find_element_by_ios_class_chain(data["hanbiro_talk_ios"]["select_contact_org"])
        if  select_contact_org.is_displayed():
            Logging("=> No Crash App") 
            TestCase_LogResult(**data["testcase_result"]["hanbiro_talk_ios"]["search_select_user"]["pass"])
            time.sleep(3)
            select_contact_org=driver.find_element_by_ios_class_chain(data["hanbiro_talk_ios"]["select_contact_org"])
            select_contact_org.click()
            time.sleep(1)
            Logging("3. Select User  successfully")
        else:
            ValidateFailResultAndSystem("=> Crash App")
            TestCase_LogResult(**data["testcase_result"]["hanbiro_talk_ios"]["search_select_user"]["fail"])
            exit(0)
    except WebDriverException:
        ValidateFailResultAndSystem("=> Crash App")
        exit(0)
    try:
        show_icon_chat_mess = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["hanbiro_talk_ios"]["icon_chat"])))
        if  show_icon_chat_mess.is_displayed():
            Logging("=> No Crash App") 
        else:
            ValidateFailResultAndSystem("=> Crash App")
            exit(0)
    except WebDriverException:
        ValidateFailResultAndSystem("=> Crash App")
        exit(0)
    time.sleep(1)
    click_icon_close_avatar = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, data["hanbiro_talk_ios"]["icon_close_view_avatar"])))
    click_icon_close_avatar.click()
    time.sleep(1)

def hanbiro_talk_ios_message_input_msg():
    Logging("------------------------------------------------------Tab Message------------------------------------------------------")
    select_tab_message=driver.find_element_by_ios_class_chain(data["hanbiro_talk_ios"]["icon_tab_message"])
    select_tab_message.click()
    time.sleep(2)
    try:
        show_icon_importtant_mess = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["hanbiro_talk_ios"]["icon_important_message"])))
        if  show_icon_importtant_mess.is_displayed():
            Logging("=> No Crash App") 
            Logging("=> 1. Access page Message successfully") 
            TestCase_LogResult(**data["testcase_result"]["hanbiro_talk_ios"]["access_tab_message"]["pass"])
        else:
            ValidateFailResultAndSystem("=> Crash App")
            TestCase_LogResult(**data["testcase_result"]["hanbiro_talk_ios"]["access_tab_message"]["fail"])
            exit(0)
    except WebDriverException:
        ValidateFailResultAndSystem("=> Crash App")
        exit(0)
    time.sleep(1)
    try:
        txt_search_contact_org123= driver.find_element_by_ios_class_chain(data["hanbiro_talk_ios"]["txt_input_search_contact_msg"])
        txt_search_contact_org123.click()
        time.sleep(2)
        input_search_contact_msg=driver.find_element_by_ios_class_chain(data["hanbiro_talk_ios"]["txt_input_search_contact_msg"])
        input_search_contact_msg.click()
        input_search_contact_msg.send_keys(data["hanbiro_talk_ios"]["search_contact_org_talk"]+"\n")
        Logging("Input user => pass")
        time.sleep(2)
    except WebDriverException:
        ValidateFailResultAndSystem("=> Crash App")
        exit(0)
    time.sleep(2)
    Logging("-----------------------Select Contact------------------------")   
    try:
        select_contact_org=driver.find_element_by_ios_class_chain(data["hanbiro_talk_ios"]["select_contact_org"])
        select_contact_org.click()
        time.sleep(3)
        show_icon_chat_mess = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["hanbiro_talk_ios"]["icon_chat"])))
        show_icon_chat_mess.click()
        time.sleep(5)
        Logging("3. Select User  successfully")
    except WebDriverException:
        ValidateFailResultAndSystem("=> Crash App")
        exit(0)
    
    Logging("-----------------------------------------------Tab Message - Write content------------------------------------------------")
    try:
        check_show_message_in = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, data["hanbiro_talk_ios"]["icon_send_content_message"])))
        if  check_show_message_in.is_displayed():
            Logging("=> No Crash App") 
            Logging("=> 1. Access screen Message successfully")
        else:
            ValidateFailResultAndSystem("=> Crash App")
            exit(0)
    except WebDriverException:
        ValidateFailResultAndSystem("=> Crash App")
        exit(0)
    time.sleep(5)

    try:
        txt_input_message = WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, data["hanbiro_talk_ios"]["txt_input_content_message"])))
        txt_input_message.click()
        txt_input_message = WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, data["hanbiro_talk_ios"]["txt_input_content_message"])))
        txt_input_message.click()
        txt_input_message.send_keys(data["hanbiro_talk_ios"]["content_message"])
        Logging("2. Input Content  successfully")
        icon_send_mesege = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, data["hanbiro_talk_ios"]["icon_send_content_message"])))
        icon_send_mesege.click()
        Logging("3. Click icon Content  successfully")    
    except WebDriverException:
        ValidateFailResultAndSystem("=> Crash App")
        exit(0)

    time.sleep(1)
    try:
        check_show_message_in = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, data["hanbiro_talk_ios"]["icon_send_content_message"])))
        if  check_show_message_in.is_displayed():
            Logging("=> No Crash App") 
            Logging("=>3. Sent content successfully")
            TestCase_LogResult(**data["testcase_result"]["hanbiro_talk_ios"]["sent_content_message"]["pass"])
        else:
            ValidateFailResultAndSystem("=> Crash App")
            TestCase_LogResult(**data["testcase_result"]["hanbiro_talk_ios"]["sent_content_message"]["fail"])
            exit(0)
    except WebDriverException:
        ValidateFailResultAndSystem("=> Crash App")
        exit(0)

    time.sleep(1)
    




    '''
    Logging("-----------------------------------------------Tab Message - View Message Old------------------------------------------------")
    time.sleep(2)
    driver.swipe(start_x=523, start_y=1089, end_x=523, end_y=1778, duration=800)
    Logging("1. Scroll to Settings successfully")
    time.sleep(2)
    driver.swipe(start_x=523, start_y=1089, end_x=523, end_y=1778, duration=800)
    Logging("1. Scroll to Settings successfully")
    try:
        check_data_message_old = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["hanbiro_talk_ios"]["check_data_search_message"])))
        if  check_data_message_old.is_displayed():
            Logging("=> No Crash App") 
            Logging("=> Show data Old successfully")
        else:
            ValidateFailResultAndSystem("=> Not show Show data Old")
            exit(0)
    except WebDriverException:
        ValidateFailResultAndSystem("=> Crash App")
        exit(0)
    time.sleep(1)
    

    Logging("-----------------------------------------------Tab Message - Search content------------------------------------------------")
    try:
        click_icon_invite_contact = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, data["hanbiro_talk_ios"]["icon_invite_contact"])))
        if  click_icon_invite_contact.is_displayed():
            Logging("=> No Crash App") 
            click_icon_invite_contact = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, data["hanbiro_talk_ios"]["icon_invite_contact"])))
            click_icon_invite_contact.click()
            Logging("=>1. Click Invite successfully")
        else:
            ValidateFailResultAndSystem("=> Crash App")
            exit(0)
    except WebDriverException:
        ValidateFailResultAndSystem("=> Crash App")
        exit(0)
    try:
        click_search_messages = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, data["hanbiro_talk_ios"]["icon_search_messages"])))
        click_search_messages.click()
        Logging("=>2. Click Icon Search messages successfully")
    except WebDriverException:
        Logging("=>2. Click Icon Search messages Fail")
        exit(0)
    try:
        txt_search_keywork_message = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, data["hanbiro_talk_ios"]["txt_search_messages"])))
        if  txt_search_keywork_message.is_displayed():
            Logging("=> No Crash App") 
            txt_search_keywork_message = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, data["hanbiro_talk_ios"]["txt_search_messages"])))
            txt_search_keywork_message.click()
            Logging("=>1. Click textbox search successfully")
            driver.is_keyboard_shown()
            driver.press_keycode(36)
            driver.press_keycode(33)
            driver.press_keycode(40)
            driver.press_keycode(40)
            driver.press_keycode(43)
            driver.press_keycode(62)
            driver.press_keycode(48)
            driver.press_keycode(33)
            driver.press_keycode(29)
            driver.press_keycode(41)
            driver.press_keycode(66)
            Logging("2. Input Content Search successfully")
            btn_search_message = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, data["hanbiro_talk_ios"]["icon_search_content"])))
            btn_search_message.click()
            check_data_search_message = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["hanbiro_talk_ios"]["check_data_search_message"])))
            if  check_data_search_message.is_displayed():
                Logging("=> No Crash App") 
                TestCase_LogResult(**data["testcase_result"]["hanbiro_talk_ios"]["search_content_message"]["pass"])
            else:
                ValidateFailResultAndSystem("=> Crash App")
                TestCase_LogResult(**data["testcase_result"]["hanbiro_talk_ios"]["search_content_message"]["fail"])
                exit(0)
            Logging("3.Search content successfully")
        else:
            ValidateFailResultAndSystem("=> Crash App")
            exit(0)
    except WebDriverException:
        ValidateFailResultAndSystem("=> Search content Fail")
        exit(0)
    Logging("-----------------------------------------------Tab Message - Filter------------------------------------------------")
    try:
        click_list_messages_filter = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, data["hanbiro_talk_ios"]["icon_filter_content"])))
        click_list_messages_filter.click()
        Logging("=>1. Click Icon Filter messages successfully")
    except WebDriverException:
        Logging("=>1. Click Icon Filter Fail")
        exit(0)
    try:
        btn_today_filter = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, data["hanbiro_talk_ios"]["btn_today_filter_content"])))
        if  btn_today_filter.is_displayed():
            Logging("=> No Crash App") 
            btn_today_filter = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, data["hanbiro_talk_ios"]["btn_today_filter_content"])))
            btn_today_filter.click()
            Logging("=>2. Click button Today successfully")
            click_icon_confirm_filter = WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.ID, data["hanbiro_talk_ios"]["icon_confirm_filter_content"])))
            click_icon_confirm_filter.click()
            Logging("3.Search Filter successfully")
            TestCase_LogResult(**data["testcase_result"]["hanbiro_talk_ios"]["filter_content_message"]["pass"])
        else:
            ValidateFailResultAndSystem("=> Crash App")
            TestCase_LogResult(**data["testcase_result"]["hanbiro_talk_ios"]["filter_content_message"]["fail"])
            exit(0)
    except WebDriverException:
        ValidateFailResultAndSystem("=> Crash App")
        exit(0)
    driver.back()
    time.sleep(1)
    '''



def hanbiro_talk_ios_message_attach_file():
    Logging("------------------------------------------------------Tab Message - Attach file------------------------------------------------------")
    icon_attach_file_image = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, data["hanbiro_talk_ios"]["icon_attach_file_image"])))
    icon_attach_file_image.click()
    time.sleep(3)
    try:
        check_crash_click_icon_attach_file = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["hanbiro_talk_ios"]["click_checkbox_select_file_image"])))
        if  check_crash_click_icon_attach_file.is_displayed():
            Logging("=> No Crash App") 
            time.sleep(1)
            Logging("=> Click Icon Attach File successfully")
        else:
            ValidateFailResultAndSystem("=> Crash App")
            exit(0)
    except WebDriverException:
        ValidateFailResultAndSystem("=> Crash App")
        exit(0)
    checkbox_select_file_image = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["hanbiro_talk_ios"]["click_checkbox_select_file_image"])))
    checkbox_select_file_image.click()
    try:
        check_crash_click_icon_attach_file = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, data["hanbiro_talk_ios"]["icon_send_content_message"])))
        if  check_crash_click_icon_attach_file.is_displayed():
            Logging("=> No Crash App") 
            Logging("=> Select File Image successfully")
            time.sleep(1)
            check_crash_click_icon_attach_file = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, data["hanbiro_talk_ios"]["icon_send_content_message"])))
            check_crash_click_icon_attach_file.click()
        else:
            ValidateFailResultAndSystem("=> Crash App")
            exit(0)
    except WebDriverException:
        ValidateFailResultAndSystem("=> Crash App")
        exit(0)
    
    time.sleep(4)

    check_crash_sent_attach_file = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["hanbiro_talk_ios"]["click_image_view_message"])))
    if  check_crash_sent_attach_file.is_displayed():
        Logging("=> No Crash App") 
        Logging("=> Sent File Image successfully")
        TestCase_LogResult(**data["testcase_result"]["hanbiro_talk_ios"]["send_file_message"]["pass"])
    else:
        ValidateFailResultAndSystem("=> Crash App")
        TestCase_LogResult(**data["testcase_result"]["hanbiro_talk_ios"]["send_file_message"]["fail"])
        exit(0)

    '''

    try:
        check_crash_sent_attach_file = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["hanbiro_talk_ios"]["click_image_view_message"])))
        if  check_crash_sent_attach_file.is_displayed():
            Logging("=> No Crash App") 
            Logging("=> Sent File Image successfully")
            TestCase_LogResult(**data["testcase_result"]["hanbiro_talk_ios"]["send_file_message"]["pass"])
        else:
            ValidateFailResultAndSystem("=> Crash App")
            TestCase_LogResult(**data["testcase_result"]["hanbiro_talk_ios"]["send_file_message"]["fail"])
            exit(0)
    except WebDriverException:
        ValidateFailResultAndSystem("=> Attach file Fail")
        exit(0)
    '''
    time.sleep(2)



    hide_sent_file = driver.find_element_by_ios_class_chain(data["hanbiro_talk_ios"]["hide_sent_image"])
    hide_sent_file.click()
    time.sleep(1)

    Logging("-----------------------------------------------Tab Message - View file------------------------------------------------")

    #par = driver.find_element_by_ios_class_chain(data["hanbiro_talk_ios"]["kkk_click_checkbox_select_file_image"])

    #par = driver.find_element_by_xpath("//XCUIElementTypeApplication[@name='HanbiroTalk']/XCUIElementTypeWindow[1]/XCUIElementTypeOther/XCUIElementTypeOther/XCUIElementTypeOther/XCUIElementTypeOther/XCUIElementTypeOther/XCUIElementTypeOther/XCUIElementTypeOther/XCUIElementTypeOther/XCUIElementTypeOther/XCUIElementTypeOther/XCUIElementTypeOther/XCUIElementTypeTables")
    par = driver.find_element_by_ios_predicate('type == "XCUIElementTypeTable"')
    Logging("=> OK OK OK ")
    time.sleep(5)
    img_nu = par.find_elements_by_xpath("//XCUIElementTypeCell")
    print("=> OK OK OK  nmnm",len(img_nu))
    i = len(img_nu)
    time.sleep(3)
    while i > 1 :
        time.sleep(3)
        m = "//XCUIElementTypeApplication[@name='HanbiroTalk']/XCUIElementTypeWindow[1]/XCUIElementTypeOther/XCUIElementTypeOther/XCUIElementTypeOther/XCUIElementTypeOther/XCUIElementTypeOther/XCUIElementTypeOther/XCUIElementTypeOther/XCUIElementTypeOther/XCUIElementTypeOther/XCUIElementTypeOther/XCUIElementTypeOther/XCUIElementTypeTable/XCUIElementTypeCell["

        m1 = str(i) + "]/XCUIElementTypeOther"
        l= m+m1
        driver.find_element_by_xpath(l).click()    
        time.sleep(1)
        print("i",i)
        i = i - 1
        break



    Logging("=> View Image successfully")
    #time.sleep(2)
    '''
    try:
        view_image_in_message = driver.find_element_by_ios_class_chain(data["hanbiro_talk_ios"]["click_icon_download_image_in_message"])
        if  view_image_in_message.is_displayed():
            Logging("=> No Crash App") 
            Logging("=> View File Image successfully")
            TestCase_LogResult(**data["testcase_result"]["hanbiro_talk_ios"]["view_file_message"]["pass"])
        else:
            ValidateFailResultAndSystem("=> Crash App")
            TestCase_LogResult(**data["testcase_result"]["hanbiro_talk_ios"]["view_file_message"]["fail"])
            exit(0)
    except WebDriverException:
        ValidateFailResultAndSystem("=> Crash App")
        exit(0)
    '''

    Logging("------------------------------------------------------Tab Message - Download file------------------------------------------------------")
    #click_download_image_in_message = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, data["hanbiro_talk_ios"]["click_icon_download_image_in_message"])))
    
    click_download_image_in_message=driver.find_element_by_ios_class_chain(data["hanbiro_talk_ios"]["click_icon_download_image_in_message"])
    click_download_image_in_message.click()
    time.sleep(1)
    Logging("=> Click icon Download file successfully")



    '''
    try:
        view_image_in_message = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, data["hanbiro_talk_ios"]["click_icon_download_image_in_message"])))
        if  view_image_in_message.is_displayed():
            Logging("=> No Crash App") 
            TestCase_LogResult(**data["testcase_result"]["hanbiro_talk_ios"]["download_file_message"]["pass"])
            #Logging("=> View File Image successfully")
        else:
            ValidateFailResultAndSystem("=> Crash App")
            TestCase_LogResult(**data["testcase_result"]["hanbiro_talk_ios"]["download_file_message"]["fail"])
            exit(0)
    except WebDriverException:
        ValidateFailResultAndSystem("=> Crash App")
        exit(0)
    '''
    




    home_img = driver.find_element_by_ios_class_chain(data["hanbiro_talk_ios"]["show_done_view_image_in_message"])
    home_img.click()
    time.sleep(1)
    click_done_view_image=driver.find_element_by_ios_class_chain(data["hanbiro_talk_ios"]["done_view_image_in_message"])
    click_done_view_image.click()
    time.sleep(2)
    



def hanbiro_talk_android_message_fw_copy():
    Logging("------------------------------------------------------Tab Message - Forward File------------------------------------------------------")

    par = driver.find_element_by_ios_predicate('type == "XCUIElementTypeTable"')
    Logging("=> OK OK OK ")
    time.sleep(5)
    img_nu = par.find_elements_by_xpath("//XCUIElementTypeCell")
    print("=> OK OK OK  nmnm",len(img_nu))
    i = len(img_nu)
    time.sleep(3)
    while i > 1 :
        time.sleep(3)
        m = "//XCUIElementTypeApplication[@name='HanbiroTalk']/XCUIElementTypeWindow[1]/XCUIElementTypeOther/XCUIElementTypeOther/XCUIElementTypeOther/XCUIElementTypeOther/XCUIElementTypeOther/XCUIElementTypeOther/XCUIElementTypeOther/XCUIElementTypeOther/XCUIElementTypeOther/XCUIElementTypeOther/XCUIElementTypeOther/XCUIElementTypeTable/XCUIElementTypeCell["

        m1 = str(i) + "]/XCUIElementTypeOther"
        l= m+m1
        huhuhu = driver.find_element_by_xpath(l)
        time.sleep(3)
        print("i",i)
        i = i - 1
        break

    time.sleep(1)
    actions = TouchAction(driver)
    actions.long_press(huhuhu)
    actions.release()
    actions.perform()
    time.sleep(1)
    Logging("=>Press and hold the image successfully")
    try:
        show_value_copy_forward = driver.find_element_by_ios_class_chain(data["hanbiro_talk_ios"]["show_content_copy_forward_in_message"])
        if  show_value_copy_forward.is_displayed():
            Logging("=> No Crash App") 
            #Logging("=> View File Image successfully")
        else:
            ValidateFailResultAndSystem("=> Crash App")
            exit(0)
    except WebDriverException:
        ValidateFailResultAndSystem("=> Crash App")
        exit(0)
    click_forward_image_in_message = driver.find_element_by_ios_class_chain(data["hanbiro_talk_ios"]["show_content_copy_forward_in_message"])
    click_forward_image_in_message.click()
    Logging("=> Click Value Forward successfully")

    time.sleep(1)
    txt_search_contact_org_fw = driver.find_element_by_ios_class_chain(data["hanbiro_talk_ios"]["txt_search_contact_org_forward"])
    if  txt_search_contact_org_fw.is_displayed():
        Logging("=> No Crash App")        
        txt_search_contact_org_fw= driver.find_element_by_ios_class_chain(data["hanbiro_talk_ios"]["txt_search_contact_org_forward"])
        txt_search_contact_org_fw.click()
        time.sleep(1)
        txt_search_contact_org_fw.send_keys(data["hanbiro_talk_ios"]["search_contact_org_forward"]+"\n")
        Logging("Input user => pass")
        time.sleep(4)
    else:
            ValidateFailResultAndSystem("=> Crash App")
            exit(0)

   
    try:
        select_contact_org=driver.find_element_by_ios_class_chain(data["hanbiro_talk_ios"]["select_contact_fw_to"])
        select_contact_org.click()
        time.sleep(2)
    except WebDriverException:
        ValidateFailResultAndSystem("=> Crash App")
        exit(0)
    time.sleep(1)
    click_send_fw_file_message = driver.find_element_by_ios_class_chain(data["hanbiro_talk_ios"]["send_fw_file_message"])
    click_send_fw_file_message.click()
    time.sleep(1)
    Logging("=> Click SEND Forward successfully")
    time.sleep(3)
    
    try:
        check_show_message_in = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, data["hanbiro_talk_ios"]["icon_attach_file_image"])))
        if  check_show_message_in.is_displayed():
            Logging("=> No Crash App") 
            Logging("=>3. Sent content successfully")
            TestCase_LogResult(**data["testcase_result"]["hanbiro_talk_ios"]["forward_file_message"]["pass"])
        else:
            ValidateFailResultAndSystem("=> Crash App")
            TestCase_LogResult(**data["testcase_result"]["hanbiro_talk_ios"]["forward_file_message"]["fail"])
            exit(0)
    except WebDriverException:
        ValidateFailResultAndSystem("=> Crash App")
        exit(0)
    time.sleep(1)
    


    
    Logging("------------------------------------------------------Tab Message - Copy------------------------------------------------------")
    copy = driver.find_element_by_ios_predicate('type == "XCUIElementTypeTable"')
    Logging("=> OK OK OK ")
    time.sleep(5)
    img_text = copy.find_elements_by_xpath("//XCUIElementTypeCell")
    print("=> OK OK OK  nmnm",len(img_text))
    i = len(img_text)
    time.sleep(3)
    while i > 1 :
        time.sleep(3)
        m = "//XCUIElementTypeApplication[@name='HanbiroTalk']/XCUIElementTypeWindow[1]/XCUIElementTypeOther/XCUIElementTypeOther/XCUIElementTypeOther/XCUIElementTypeOther/XCUIElementTypeOther/XCUIElementTypeOther/XCUIElementTypeOther/XCUIElementTypeOther/XCUIElementTypeOther/XCUIElementTypeOther/XCUIElementTypeOther/XCUIElementTypeTable/XCUIElementTypeCell["

        m1 = str(i) + "]/XCUIElementTypeOther"
        l= m+m1
        hihihi = driver.find_element_by_xpath(l)
        time.sleep(3)
        print("i",i)
        i = i - 1
        break

    time.sleep(1)
    actions = TouchAction(driver)
    actions.long_press(hihihi)
    actions.release()
    actions.perform()
    time.sleep(1)
    Logging("=>Press and hold the image successfully")
    try:
        show_value_copy_forward = driver.find_element_by_ios_class_chain(data["hanbiro_talk_ios"]["show_content_copy_forward_in_message"])
        if  show_value_copy_forward.is_displayed():
            Logging("=> No Crash App") 
            #Logging("=> View File Image successfully")
        else:
            ValidateFailResultAndSystem("=> Crash App")
            exit(0)
    except WebDriverException:
        ValidateFailResultAndSystem("=> Crash App")
        exit(0)

    click_copy_image_in_message = driver.find_element_by_ios_class_chain(data["hanbiro_talk_ios"]["select_value_copy_in_message"])
    click_copy_image_in_message.click()
    Logging("=> Click Value Copy successfully")
    time.sleep(3)

    txt_input_message = WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, data["hanbiro_talk_ios"]["txt_input_content_message"])))
    txt_input_message.click()
    txt_input_message = WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, data["hanbiro_talk_ios"]["txt_input_content_message"])))
    txt_input_message.click()
    time.sleep(1)
    click_paste_in_message = driver.find_element_by_ios_class_chain(data["hanbiro_talk_ios"]["select_value_paste_in_message"])
    click_paste_in_message.click()
    Logging("=> Click Value Paste successfully")
    time.sleep(2)
    Logging("2. Input Content  successfully")
    icon_send_mesege = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, data["hanbiro_talk_ios"]["icon_send_content_message"])))
    icon_send_mesege.click()
    Logging("3. Click icon Content  successfully")  

    Logging(" =>Copy/Paste successfully")
    try:
        check_show_message_in = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, data["hanbiro_talk_ios"]["icon_attach_file_image"])))
        if  check_show_message_in.is_displayed():
            Logging("=> No Crash App") 
            Logging("=>3. Sent content successfully")
            TestCase_LogResult(**data["testcase_result"]["hanbiro_talk_ios"]["copy_message"]["pass"])
        else:
            ValidateFailResultAndSystem("=> Crash App")
            TestCase_LogResult(**data["testcase_result"]["hanbiro_talk_ios"]["copy_message"]["fail"])
            exit(0)
    except WebDriverException:
        ValidateFailResultAndSystem("=> Crash App")
        exit(0)
    time.sleep(1)









    
    
    Logging("-----------------------------------------------Tab Message - Quote message------------------------------------------------")
    

    quote = driver.find_element_by_ios_predicate('type == "XCUIElementTypeTable"')
    Logging("=> OK OK OK ")
    time.sleep(5)
    img_text_quote = quote.find_elements_by_xpath("//XCUIElementTypeCell")
    print("=> OK OK OK  nmnm",len(img_text_quote))
    i = len(img_text_quote)
    time.sleep(3)
    while i > 1 :
        time.sleep(3)
        k = "//XCUIElementTypeApplication[@name='HanbiroTalk']/XCUIElementTypeWindow[1]/XCUIElementTypeOther/XCUIElementTypeOther/XCUIElementTypeOther/XCUIElementTypeOther/XCUIElementTypeOther/XCUIElementTypeOther/XCUIElementTypeOther/XCUIElementTypeOther/XCUIElementTypeOther/XCUIElementTypeOther/XCUIElementTypeOther/XCUIElementTypeTable/XCUIElementTypeCell["

        k1 = str(i) + "]/XCUIElementTypeTextView"
        l= k+k1
        quoquo = driver.find_element_by_xpath(l)
        time.sleep(3)
        print("i",i)
        i = i - 1
        break

    time.sleep(1)
    actions = TouchAction(driver)
    actions.long_press(quoquo)
    actions.release()
    actions.perform()
    time.sleep(1)



    try:
        show_value_copy_forward = driver.find_element_by_ios_class_chain(data["hanbiro_talk_ios"]["show_content_copy_forward_in_message"])
        if  show_value_copy_forward.is_displayed():
            Logging("=> No Crash App") 
            #Logging("=> View File Image successfully")
        else:
            ValidateFailResultAndSystem("=> Crash App")
            exit(0)
    except WebDriverException:
        ValidateFailResultAndSystem("=> Crash App")
        exit(0)


    click_quote_image_in_message = driver.find_element_by_ios_class_chain(data["hanbiro_talk_ios"]["select_value_quote_in_message"])
    click_quote_image_in_message.click()
    Logging("=> Click Value Quote successfully")
    time.sleep(3)

    icon_send_mesege = WebDriverWait(driver, 50).until(EC.presence_of_element_located((By.ID, data["hanbiro_talk_ios"]["icon_send_content_message"])))
    icon_send_mesege.click()
    Logging(" =>Quote message successfully")
    time.sleep(1)
    '''

    try:
        check_show_message_in = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, data["hanbiro_talk_ios"]["icon_attach_file_image"])))
        if  check_show_message_in.is_displayed():
            Logging("=> No Crash App") 
            Logging("=>3. Sent content successfully")
            TestCase_LogResult(**data["testcase_result"]["hanbiro_talk_ios"]["copy_message"]["pass"])
        else:
            ValidateFailResultAndSystem("=> Crash App")
            TestCase_LogResult(**data["testcase_result"]["hanbiro_talk_ios"]["copy_message"]["fail"])
            exit(0)
    except WebDriverException:
        ValidateFailResultAndSystem("=> Crash App")
        exit(0)
    time.sleep(1)
    '''



    Logging("-----------------------------------------------Tab Message - Important message------------------------------------------------")
    time.sleep(2)
    important = driver.find_element_by_ios_predicate('type == "XCUIElementTypeTable"')
    Logging("=> OK OK OK ")
    time.sleep(5)
    important_mes = important.find_elements_by_xpath("//XCUIElementTypeCell")
    print("=> OK OK OK  nmnm",len(important_mes))
    i = len(important_mes)
    time.sleep(3)
    while i > 1 :
        time.sleep(3)
        m = "//XCUIElementTypeApplication[@name='HanbiroTalk']/XCUIElementTypeWindow[1]/XCUIElementTypeOther/XCUIElementTypeOther/XCUIElementTypeOther/XCUIElementTypeOther/XCUIElementTypeOther/XCUIElementTypeOther/XCUIElementTypeOther/XCUIElementTypeOther/XCUIElementTypeOther/XCUIElementTypeOther/XCUIElementTypeOther/XCUIElementTypeTable/XCUIElementTypeCell["

        m1 = str(i) + "]/XCUIElementTypeOther"
        l= m+m1
        hohoho = driver.find_element_by_xpath(l)
        time.sleep(3)
        print("i",i)
        i = i - 1
        break

    time.sleep(1)
    actions = TouchAction(driver)
    actions.long_press(hohoho)
    actions.release()
    actions.perform()
    time.sleep(1)
    Logging("=>Press and hold the image successfully")
    try:
        show_value_copy_forward = driver.find_element_by_ios_class_chain(data["hanbiro_talk_ios"]["show_content_copy_forward_in_message"])
        if  show_value_copy_forward.is_displayed():
            Logging("=> No Crash App") 
            #Logging("=> View File Image successfully")
        else:
            ValidateFailResultAndSystem("=> Crash App")
            exit(0)
    except WebDriverException:
        ValidateFailResultAndSystem("=> Crash App")
        exit(0)
    click_copy_image_in_message = driver.find_element_by_ios_class_chain(data["hanbiro_talk_ios"]["select_value_important_in_message"])
    click_copy_image_in_message.click()
    Logging("=> Click Value Important message successfully")
    time.sleep(3)


    '''
    txt_input_message = WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, data["hanbiro_talk_ios"]["txt_input_content_message"])))
    txt_input_message.click()
    txt_input_message = WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, data["hanbiro_talk_ios"]["txt_input_content_message"])))
    txt_input_message.click()
    time.sleep(1)
    click_paste_in_message = driver.find_element_by_ios_class_chain(data["hanbiro_talk_ios"]["select_value_paste_in_message"])
    click_paste_in_message.click()
    Logging("=> Click Value Paste successfully")
    time.sleep(2)
    Logging("2. Input Content  successfully")
    icon_send_mesege = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, data["hanbiro_talk_ios"]["icon_send_content_message"])))
    icon_send_mesege.click()
    Logging("3. Click icon Content  successfully")  
    '''
    Logging(" =>Important message successfully")
    try:
        check_show_message_in = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, data["hanbiro_talk_ios"]["icon_attach_file_image"])))
        if  check_show_message_in.is_displayed():
            Logging("=> No Crash App") 
            Logging("=>3. Sent content successfully")
            TestCase_LogResult(**data["testcase_result"]["hanbiro_talk_ios"]["copy_message"]["pass"])
        else:
            ValidateFailResultAndSystem("=> Crash App")
            TestCase_LogResult(**data["testcase_result"]["hanbiro_talk_ios"]["copy_message"]["fail"])
            exit(0)
    except WebDriverException:
        ValidateFailResultAndSystem("=> Crash App")
        exit(0)
    time.sleep(1)




    Logging("-----------------------------------------------Tab Message - View file list------------------------------------------------")
    try:
        click_icon_more_contact = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["hanbiro_talk_ios"]["icon_more_contact"])))
        if  click_icon_more_contact.is_displayed():
            Logging("=> No Crash App") 
            time.sleep(1)
            click_icon_more_contact = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["hanbiro_talk_ios"]["icon_more_contact"])))
            click_icon_more_contact.click()
            Logging("=>1. Click More successfully")
        else:
            ValidateFailResultAndSystem("=> Crash App")
            exit(0)
    except WebDriverException:
        ValidateFailResultAndSystem("=> Crash App")
        exit(0)
    try:
        click_view_file_list = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["hanbiro_talk_ios"]["click_file_list"])))
        click_view_file_list.click()
        time.sleep(1)
        show_photo_in_file_list = WebDriverWait(driver, 50).until(EC.presence_of_element_located((By.XPATH, data["hanbiro_talk_ios"]["view_photo_in_file_list"])))
        if  show_photo_in_file_list.is_displayed():
            Logging("=> show image in file list") 
            TestCase_LogResult(**data["testcase_result"]["hanbiro_talk_ios"]["view_file_list_message"]["pass"])
        else:
            ValidateFailResultAndSystem("=> Not show image in file list")
            TestCase_LogResult(**data["testcase_result"]["hanbiro_talk_ios"]["view_file_list_message"]["fail"])
            exit(0)
    except WebDriverException:
        ValidateFailResultAndSystem("=> Crash App")
        exit(0)






   
    driver.back()
    time.sleep(3)
    driver.swipe(start_x=151, start_y=460, end_x=523, end_y=157, duration=543)
    time.sleep(5)

    
    
    Logging("------------------------------------------------------Tab Message - Add new member------------------------------------------------------")


    try:
        click_add_new_member = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["hanbiro_talk_ios"]["click_add_new_member"])))
        click_add_new_member.click()
    except WebDriverException:
        ValidateFailResultAndSystem("=> Crash App")
        exit(0)
    
    txt_search_contact_org_add_member = driver.find_element_by_ios_class_chain(data["hanbiro_talk_ios"]["txt_search_contact_org_add_memeber"])
    if  txt_search_contact_org_add_member.is_displayed():
        Logging("=> No Crash App") 
        txt_search_contact_org_add_member.click()
        time.sleep(3)       
        txt_search_contact_org_add_member= driver.find_element_by_ios_class_chain(data["hanbiro_talk_ios"]["txt_search_contact_org_add_memeber"])
        txt_search_contact_org_add_member.click()
        time.sleep(1)
        txt_search_contact_org_add_member.send_keys(data["hanbiro_talk_ios"]["search_contact_org_add_member"]+"\n")
        Logging("Input user => pass")
        time.sleep(4)
    else:
            ValidateFailResultAndSystem("=> Crash App")
            exit(0)

    time.sleep(4)
    try:
        select_contact_org=driver.find_element_by_ios_class_chain(data["hanbiro_talk_ios"]["select_contact_add_memeber"])
        select_contact_org.click()
        time.sleep(2)
        confirm_elect_contact_org=driver.find_element_by_ios_class_chain(data["hanbiro_talk_ios"]["conffirm_add_contact_org_add_member"])
        confirm_elect_contact_org.click()
        time.sleep(2)
        Logging("Add Contact => pass")
    except WebDriverException:
        ValidateFailResultAndSystem("=> Crash App")
        exit(0)
    time.sleep(100)





    driver.swipe(start_x=523, start_y=1778, end_x=523, end_y=1089, duration=800)
    time.sleep(1)


    '''
    select_add_number= WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["hanbiro_talk_ios"]["select_add_new_member"])))
    select_add_number.click()    
    try:
        txt_search_contact_org = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, data["hanbiro_talk_ios"]["txt_search_contact"])))
        if  txt_search_contact_org.is_displayed():
            Logging("=> No Crash App") 
            txt_search_contact_org = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, data["hanbiro_talk_ios"]["txt_search_contact"])))
            txt_search_contact_org.click()
            driver.is_keyboard_shown()
            driver.press_keycode(29)
            driver.press_keycode(49)
            driver.press_keycode(48)
            driver.press_keycode(43)
            driver.press_keycode(41)
            driver.press_keycode(29)
            driver.press_keycode(66)
            Logging("2. Input Search User  successfully")
        else:
            ValidateFailResultAndSystem("=> Crash App")
            exit(0)
    except WebDriverException:
        ValidateFailResultAndSystem("=> Crash App")
        exit(0)
    try:
        select_contact_org = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["hanbiro_talk_ios"]["select_contact_add_number"])))
        if  select_contact_org.is_displayed():
            Logging("=> No Crash App") 
            select_contact_org = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["hanbiro_talk_ios"]["select_contact_add_number"])))
            select_contact_org.click()
            time.sleep(1)
            Logging("3. Select User  successfully")
            TestCase_LogResult(**data["testcase_result"]["hanbiro_talk_ios"]["add_member"]["pass"])
        else:
            ValidateFailResultAndSystem("=> Crash App")
            TestCase_LogResult(**data["testcase_result"]["hanbiro_talk_ios"]["add_member"]["fail"])
            exit(0)
    except WebDriverException:
        ValidateFailResultAndSystem("=> Crash App")
        exit(0)
    click_icon_confirm_add_user = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, data["hanbiro_talk_ios"]["icon_confirm_contact_add_member"])))
    click_icon_confirm_add_user.click()
    Logging("4. Add Member  successfully")
    time.sleep(1)
    #driver.back()
    #time.sleep(1)
    #driver.back()
    #time.sleep(1)
    #driver.back()
    #time.sleep(1)





    
    
    Logging("------------------------------------------------------Tab Message - Change room title------------------------------------------------------")

    try:
        click_change_name = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["hanbiro_talk_ios"]["click_change_room_name"])))
        click_change_name.click()
        time.sleep(1)


        enter_name_rooom=driver.find_element_by_ios_class_chain(data["hanbiro_talk_ios"]["txt_input_room"])
        enter_name_rooom.click()
        time.sleep(2)
        Logging("click  => pass")
        click_change_room_title.clear()
        time.sleep(5)
        enter_name_rooom.send_keys(data["hanbiro_talk_ios"]["room_name_change"]+"\n")
        click_save_change_room_title= WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["hanbiro_talk_ios"]["save_change_room"])))
        time.sleep(1)
        click_save_change_room_title.click() 
        Logging("Input content  => pass")
        
    except WebDriverException:
        ValidateFailResultAndSystem("=> Crash App")
        exit(0)

    time.sleep(1000)
    try:
        click_change_room_title= WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["hanbiro_talk_ios"]["click_change_room_title"])))
        click_change_room_title.clear()
        time.sleep(1)
        click_change_room_title.click() 
        time.sleep(1)
        driver.press_keycode(46)
        time.sleep(1)
        driver.press_keycode(43)
        time.sleep(1)
        driver.press_keycode(43)
        time.sleep(1)
        driver.press_keycode(41)
        time.sleep(1)
        driver.press_keycode(62)
        time.sleep(1)
        driver.press_keycode(48)
        time.sleep(1)
        driver.press_keycode(33)
        time.sleep(1)
        driver.press_keycode(47)
        time.sleep(1)
        driver.press_keycode(48)
        time.sleep(1)
        click_change_room_title= WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["hanbiro_talk_ios"]["click_btn_ok_change_room"])))
        click_change_room_title.click() 
        time.sleep(1)
        driver.back()
        check_data_change_room_title = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["hanbiro_talk_ios"]["check_data_change_room"])))
        if  check_data_change_room_title.is_displayed():
            Logging("=> Change room title  successfully") 
            TestCase_LogResult(**data["testcase_result"]["hanbiro_talk_ios"]["change_room_title"]["pass"])
        else:
            Logging("=> Change room title Fail") 
            TestCase_LogResult(**data["testcase_result"]["hanbiro_talk_ios"]["change_room_title"]["fail"])
        time.sleep(2)
    except WebDriverException:
        ValidateFailResultAndSystem("=> Crash App")
    driver.back()
    time.sleep(1)
    '''
    driver.back()
    time.sleep(1)

    

def hanbiro_talk_ios_write_board():

    #check_crash_app_log_in_hanbiro_talk = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["hanbiro_talk_android"]["check_crash_app_log_in"]))).click()
    Logging("------------------------------------------------------Tab Message - Board -----------------------------------------------------")
    '''
    select_tab_message=driver.find_element_by_ios_class_chain(data["hanbiro_talk_ios"]["icon_tab_message"])
    select_tab_message.click()
    time.sleep(1)
    try:
        txt_search_contact_org123= driver.find_element_by_ios_class_chain(data["hanbiro_talk_ios"]["txt_input_search_contact_msg"])
        txt_search_contact_org123.click()
        time.sleep(2)
        input_search_contact_msg=driver.find_element_by_ios_class_chain(data["hanbiro_talk_ios"]["txt_input_search_contact_msg"])
        input_search_contact_msg.click()
        input_search_contact_msg.send_keys(data["hanbiro_talk_ios"]["search_contact_org_talk"]+"\n")
        Logging("Input user => pass")
        time.sleep(1)
    except WebDriverException:
        ValidateFailResultAndSystem("=> Crash App")
        exit(0)
    time.sleep(1)
    Logging("-----------------------Select Contact------------------------")   
    try:
        select_contact_org=driver.find_element_by_ios_class_chain(data["hanbiro_talk_ios"]["select_contact_org"])
        select_contact_org.click()
        time.sleep(1)
        show_icon_chat_mess = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["hanbiro_talk_ios"]["icon_chat"])))
        show_icon_chat_mess.click()
        time.sleep(1)
        Logging("3. Select User  successfully")
    except WebDriverException:
        ValidateFailResultAndSystem("=> Crash App")
        exit(0)
    time.sleep(1)
    '''

    
    try:
        click_icon_more_contact = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["hanbiro_talk_ios"]["icon_more_contact"])))
        if  click_icon_more_contact.is_displayed():
            Logging("=> No Crash App") 
            time.sleep(1)
            click_icon_more_contact = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["hanbiro_talk_ios"]["icon_more_contact"])))
            click_icon_more_contact.click()
            Logging("=>1. Click More successfully")
        else:
            ValidateFailResultAndSystem("=> Crash App")
            exit(0)
    except WebDriverException:
        ValidateFailResultAndSystem("=> Crash App")
        exit(0)
    
    #driver.swipe(start_x=148, start_y=167, end_x=627, end_y=82, duration=460)
    #time.sleep(5)

    try:
        show_option_board = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["hanbiro_talk_ios"]["click_board"])))
        if  show_option_board.is_displayed():
            Logging("=> No Crash App") 
            show_option_board = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["hanbiro_talk_ios"]["click_board"])))
            show_option_board.click()
            time.sleep(1)
            Logging("=> 2. Click BOARD successfully")
        else:
            ValidateFailResultAndSystem("=> Crash App")
            exit(0)
    except WebDriverException:
        ValidateFailResultAndSystem("=> Crash App")
        exit(0)

    '''
    try:
        check_show_option_board = driver.find_element_by_ios_class_chain(data["hanbiro_talk_ios"]["check_show_board_contact"])
        if  check_show_option_board.is_displayed():
            Logging("=> No Crash App") 
            Logging("=> 2. Click BOARD successfully")
        else:
            ValidateFailResultAndSystem("=> Crash App")
            exit(0)
    except WebDriverException:
        ValidateFailResultAndSystem("=> Crash App")
        exit(0)

    '''
    time.sleep(2)
    click_icon_write_board = driver.find_element_by_ios_class_chain(data["hanbiro_talk_ios"]["icon_write_board"])
    click_icon_write_board.click()
    Logging("=> Click icon Create Board Pass ")
    

    try:
        check_click_icon_write_board = driver.find_element_by_ios_class_chain(data["hanbiro_talk_ios"]["show_check_box_notifications"])
        if  check_click_icon_write_board.is_displayed():
            Logging("=> No Crash App")
            Logging("=> 3. Click Icon Write BOARD successfully") 
            check_click_icon_noti_write_board = driver.find_element_by_ios_class_chain(data["hanbiro_talk_ios"]["show_check_box_notifications"])
            check_click_icon_noti_write_board.click()
            Logging("=> 4. Check Notifications Write BOARD successfully")
        else:
            ValidateFailResultAndSystem("=> Crash App")
            exit(0)
    except WebDriverException:
        ValidateFailResultAndSystem("=> Crash App")
        exit(0)

    
    enter_message_write_board=driver.find_element_by_ios_class_chain(data["hanbiro_talk_ios"]["input_message_board"])
    enter_message_write_board.click()
    enter_message_write_board.send_keys(data["hanbiro_talk_ios"]["content_message_board"]+"\n")
    Logging("Input content  => pass")
    
   
    time.sleep(1)
    try:
        check_click_icon_write_board = driver.find_element_by_ios_class_chain(data["hanbiro_talk_ios"]["show_check_box_notifications"])
        if  check_click_icon_write_board.is_displayed():
            Logging("=> No Crash App") 
            time.sleep(1)
        else:
            ValidateFailResultAndSystem("=> Crash App")
            exit(0)
    except WebDriverException:
        ValidateFailResultAndSystem("=> Crash App")



    Logging("------------------------------------------------------Board - Attach file -----------------------------------------------------")
    
    click_icon_attach_file_write_board = driver.find_element_by_ios_class_chain(data["hanbiro_talk_ios"]["click_icon_attach_file_board"])
    click_icon_attach_file_write_board.click()
    try:
        checkbox_select_file_image = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["hanbiro_talk_ios"]["check_select_attach_file_image_board"])))
        if  checkbox_select_file_image.is_displayed():
            Logging("=> No Crash App") 
            time.sleep(1)
            Logging("=>1. Select File Image successfully")
        else:
            ValidateFailResultAndSystem("=> Crash App")
            exit(0)
    except WebDriverException:
        ValidateFailResultAndSystem("=> Crash App")
        exit(0)
    time.sleep(1)
    checkbox_select_file_image = WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.XPATH, data["hanbiro_talk_ios"]["check_select_attach_file_image_board"])))
    checkbox_select_file_image.click()

    time.sleep(1)

    try:
        check_crash_select_file_wirte_board = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["hanbiro_talk_ios"]["icon_done_send_file_board"])))
        if  check_crash_select_file_wirte_board.is_displayed():
            Logging("=> No Crash App") 
            time.sleep(1)
            done_select_image_wirte_board = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["hanbiro_talk_ios"]["icon_done_send_file_board"])))
            done_select_image_wirte_board.click()
            Logging("=> 2. ADD  File - Board successfully")
        else:
            ValidateFailResultAndSystem("=> Crash App")
            exit(0)
    except WebDriverException:
        ValidateFailResultAndSystem("=> Crash App")
        exit(0)
    
    icon_send_board = WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, data["hanbiro_talk_ios"]["icon_send_board"])))
    icon_send_board.click()
    Logging("=> Save Board successfully")
    time.sleep(5)







    '''
    try:
        check_show_option_board = WebDriverWait(driver, 50).until(EC.presence_of_element_located((By.ID, data["hanbiro_talk_ios"]["check_show_board_contact"])))
        if  check_show_option_board.is_displayed():
            Logging("=> No Crash App") 
            Logging("=> 3. Click Icon Send Board successfully") 
        else:
            ValidateFailResultAndSystem("=> Crash App")
            exit(0)
    except WebDriverException:
        ValidateFailResultAndSystem("=> Crash App")
        exit(0)
    time.sleep(1)
    driver.back()
    time.sleep(100)
    '''

    try:
        check_save_board_success = driver.find_element_by_ios_class_chain(data["hanbiro_talk_ios"]["icon_write_board"])
        if  check_save_board_success.is_displayed():
            Logging("=> No Crash App") 
            Logging("=> Send Board successfully")
            TestCase_LogResult(**data["testcase_result"]["hanbiro_talk_ios"]["write_board"]["pass"]) 
        else:
            ValidateFailResultAndSystem("=> Crash App")
            exit(0)
    except WebDriverException:
        ValidateFailResultAndSystem("=> Crash App")
        TestCase_LogResult(**data["testcase_result"]["hanbiro_talk_ios"]["write_board"]["fail"])
        exit(0)

    
def hanbiro_talk_ios_board_view_download_file():
    Logging("------------------------------------------------------Board - View File -----------------------------------------------------")
    '''
    par_board= driver.find_element_by_ios_predicate('type == "XCUIElementTypeTable"')
    Logging("=> OK OK OK ")
    time.sleep(5)
    img_num= par_board.find_elements_by_xpath("//XCUIElementTypeCell[3]")
    print("=> OK OK OK  nmnm",len(img_num))
    i = len(img_num)
    time.sleep(3)
    while i > 1 :
        time.sleep(3)
        me= "//XCUIElementTypeApplication[@name='HanbiroTalk']/XCUIElementTypeWindow[1]/XCUIElementTypeOther/XCUIElementTypeOther/XCUIElementTypeOther/XCUIElementTypeOther/XCUIElementTypeOther/XCUIElementTypeOther/XCUIElementTypeOther/XCUIElementTypeOther/XCUIElementTypeOther/XCUIElementTypeOther/XCUIElementTypeOther/XCUIElementTypeTable/XCUIElementTypeCell["

        m1 = str(i) + "]/XCUIElementTypeOther"
        lu= me+m1
        cell = driver.find_element_by_xpath(lu)
        time.sleep(3)
        print("i",i)
        i = i - 1
        break
    '''






    try:
        check_show_image_board = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["hanbiro_talk_ios"]["check_show_view_board_success"])))
        if  check_show_image_board.is_displayed():
            Logging("=> No Crash App") 
            check_show_image_board = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["hanbiro_talk_ios"]["check_show_view_board_success"])))
            check_show_image_board.click()
            Logging("=> 1. Click image in Board successfully") 
        else:
            ValidateFailResultAndSystem("=> Crash App")
            exit(0)
    except WebDriverException:
        ValidateFailResultAndSystem("=> Crash App")
        exit(0)
    time.sleep(1)

    
  


    '''

    try:
        show_image_board = driver.find_element_by_ios_class_chain(data["hanbiro_talk_ios"]["view_image_board_success"])
        if  show_image_board.is_displayed():
            Logging("=> No Crash App") 
            show_image_board = driver.find_element_by_ios_class_chain(data["hanbiro_talk_ios"]["view_image_board_success"])
            show_image_board.click()
            Logging("=> 2. View image in Board successfully") 
        else:
            ValidateFailResultAndSystem("=> Crash App")
            TestCase_LogResult(**data["testcase_result"]["hanbiro_talk_ios"]["view_file_board"]["fail"])
            exit(0)
    except WebDriverException:
        ValidateFailResultAndSystem("=> Crash App")
        exit(0)
    time.sleep(200)
    

    Logging("------------------------------------------------------Board - Download File -----------------------------------------------------")
    try:
        download_file_board = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, data["hanbiro_talk_ios"]["icon_download_file_board"])))
        if  download_file_board.is_displayed():
            Logging("=> No Crash App") 
            download_file_board = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, data["hanbiro_talk_ios"]["icon_download_file_board"])))
            download_file_board.click()
            Logging("=> 1. Download image in Board successfully") 
            TestCase_LogResult(**data["testcase_result"]["hanbiro_talk_ios"]["download_file_board"]["pass"])
        else:
            ValidateFailResultAndSystem("=> Crash App")
            TestCase_LogResult(**data["testcase_result"]["hanbiro_talk_ios"]["download_file_board"]["fail"])
            exit(0)
    except WebDriverException:
        ValidateFailResultAndSystem("=> Crash App")
        exit(0)
    try:
        download_file_board = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, data["hanbiro_talk_ios"]["icon_download_file_board"])))
        if  download_file_board.is_displayed():
            Logging("=> No Crash App") 
            Logging("=> Download image in Board successfully") 
        else:
            ValidateFailResultAndSystem("=> Crash App")
            exit(0)
    except WebDriverException:
        ValidateFailResultAndSystem("=> Crash App")
        exit(0)
    time.sleep(1)
    driver.back()

    '''

def hanbiro_talk_ios_edit_board():
    Logging("------------------------------------------------------Board - Edit Board -----------------------------------------------------")
    try:
        icon_more_board = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["hanbiro_talk_ios"]["icon_edit_board"])))
        if  icon_more_board.is_displayed():
            Logging("=> No Crash App") 
            icon_more_board = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["hanbiro_talk_ios"]["icon_edit_board"])))
            icon_more_board.click()
            Logging("=> 1. Click Icon Edit  Board successfully") 
        else:
            ValidateFailResultAndSystem("=> Crash App")
            exit(0)
    except WebDriverException:
        ValidateFailResultAndSystem("=> Crash App")
        exit(0)
    time.sleep(1)
    try:
        select_value_edit_board = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["hanbiro_talk_ios"]["select_value_edit_board"])))
        if  select_value_edit_board.is_displayed():
            Logging("=> No Crash App") 
            select_value_edit_board = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["hanbiro_talk_ios"]["select_value_edit_board"])))
            select_value_edit_board.click()
            Logging("=> 2. Click Value Edit Board successfully") 
        else:
            ValidateFailResultAndSystem("=> Crash App")
            exit(0)
    except WebDriverException:
        ValidateFailResultAndSystem("=> Crash App")
        exit(0)
    time.sleep(1)

    enter_message_edit_board=driver.find_element_by_ios_class_chain(data["hanbiro_talk_ios"]["input_message_board"])
    enter_message_edit_board.click()
    enter_message_edit_board.send_keys(data["hanbiro_talk_ios"]["edit_content_message_board"]+"\n")
    Logging("Edit Board => pass")
    icon_send_board = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["hanbiro_talk_ios"]["icon_send_board"])))
    icon_send_board.click()
    Logging("=> Edit Board successfully")
    

    try:
        icon_more_board = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["hanbiro_talk_ios"]["icon_edit_board"])))
        if  icon_more_board.is_displayed():
            Logging("=> No Crash App") 
            Logging("5. Edit Board successfully") 
            TestCase_LogResult(**data["testcase_result"]["hanbiro_talk_ios"]["edit_board"]["pass"])
        else:
            ValidateFailResultAndSystem("=> Crash App")
            TestCase_LogResult(**data["testcase_result"]["hanbiro_talk_ios"]["edit_board"]["fail"])
            exit(0)
    except WebDriverException:
        ValidateFailResultAndSystem("=> Crash App")
        exit(0)
    
def hanbiro_talk_ios_delete_board():
    Logging("------------------------------------------------------Board - Delete Board -----------------------------------------------------")
    try:
        icon_more_board = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["hanbiro_talk_ios"]["icon_edit_board"])))
        icon_more_board.click()
        time.sleep(1)
        select_value_delete_board = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["hanbiro_talk_ios"]["select_value_delete_board"])))
        select_value_delete_board.click()
        Logging("=> 2. Click Value Delete Board successfully") 
    except WebDriverException:
        ValidateFailResultAndSystem("=> Crash App")
        exit(0)
   
    
    time.sleep(1)
    driver.back()
    time.sleep(1)
    driver.back()
    time.sleep(3)



def whisper_hanbiro_talk_ios_send_whisper():
    Logging("----------------1. Whisper Hanbiro Talk-----------------------------")
    select_tab_whisper=driver.find_element_by_ios_class_chain(data["hanbiro_talk_ios"]["click_tab_whisper"])
    select_tab_whisper.click()
    time.sleep(3)

    try:
        check_show_create_new_wew_whisper = WebDriverWait(driver, 50).until(EC.presence_of_element_located((By.XPATH, data["hanbiro_talk_ios"]["click_icon_new_whisper"])))
        if  check_show_create_new_wew_whisper.is_displayed():
            Logging("=> No Crash App") 
            TestCase_LogResult(**data["testcase_result"]["hanbiro_talk_ios"]["access_tab_whisper"]["pass"])
            check_show_create_new_wew_whisper = WebDriverWait(driver, 50).until(EC.presence_of_element_located((By.XPATH, data["hanbiro_talk_ios"]["click_icon_new_whisper"])))
            check_show_create_new_wew_whisper.click()
        else:
            ValidateFailResultAndSystem("=> Crash App")
            TestCase_LogResult(**data["testcase_result"]["hanbiro_talk_ios"]["access_tab_whisper"]["fail"])
            exit(0)
    except WebDriverException:
        ValidateFailResultAndSystem("=> Crash App")
        exit(0)
    
    time.sleep(5)
    enter_content_write_wwhisper=driver.find_element_by_ios_class_chain(data["hanbiro_talk_ios"]["txt_input_content_whisper"])
    enter_content_write_wwhisper.click()
    enter_content_write_wwhisper.send_keys(data["hanbiro_talk_ios"]["content_wwhisper"]+"\n")
    Logging("Input content  => pass")

    time.sleep(2)
    Logging("1. Input Content Whisper  successfully")



    icon_add_recipient = driver.find_element_by_ios_class_chain(data["hanbiro_talk_ios"]["write_whisper_add_contact"])
    if  icon_add_recipient.is_displayed():
        Logging("=> No Crash App")        
        icon_add_recipient= driver.find_element_by_ios_class_chain(data["hanbiro_talk_ios"]["write_whisper_add_contact"])
        icon_add_recipient.click()
        time.sleep(1)

        txt_search_contact_wwhisper= WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["hanbiro_talk_ios"]["txt_search_contact"])))
        txt_search_contact_wwhisper.click()
        time.sleep(2)
        Logging("BBBB")
        input_search_contact_org=driver.find_element_by_ios_class_chain(data["hanbiro_talk_ios"]["txt_input_search_contact"])
        input_search_contact_org.click()
        input_search_contact_org.send_keys(data["hanbiro_talk_ios"]["search_contact_org_talk"]+"\n")
        Logging("Input user => pass")
        time.sleep(1)
        select_contact_org=driver.find_element_by_ios_class_chain(data["hanbiro_talk_ios"]["select_contact_org"])
        select_contact_org.click()
        time.sleep(1)
        icon_confirm_contact_recippient_whisper = WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.XPATH, data["hanbiro_talk_ios"]["icon_confirm_contact_recipient_whisper"])))
        icon_confirm_contact_recippient_whisper.click()
        Logging("=> 5. Click Icon Confirm Contact successfully") 
        time.sleep(1)
    else:
            ValidateFailResultAndSystem("=> Crash App")
            exit(0)

    



    Logging("------------------------------------------------------Whisper - Attach file -----------------------------------------------------")
    
    icon_icon_add_file_whisper = driver.find_element_by_ios_class_chain(data["hanbiro_talk_ios"]["icon_add_file_send_whisper"])
    icon_icon_add_file_whisper.click()
    time.sleep(1)
    try:
        checkbox_select_file_image = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["hanbiro_talk_ios"]["check_select_attach_file_image_board"])))
        if  checkbox_select_file_image.is_displayed():
            Logging("=> No Crash App") 
            checkbox_select_file_image = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["hanbiro_talk_ios"]["check_select_attach_file_image_board"])))
            checkbox_select_file_image.click()
            Logging("=>1. Select File Image successfully")
        else:
            ValidateFailResultAndSystem("=> Crash App")
            exit(0)
    except WebDriverException:
        ValidateFailResultAndSystem("=> Crash App")
        exit(0)
    time.sleep(1)
    
    try:
        check_crash_select_file_wirte_whisper = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["hanbiro_talk_ios"]["icon_done_send_file_board"])))
        if  check_crash_select_file_wirte_whisper.is_displayed():
            Logging("=> No Crash App") 
            time.sleep(1)
            done_select_image_wirte_wwhisper= WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["hanbiro_talk_ios"]["icon_done_send_file_board"])))
            done_select_image_wirte_wwhisper.click()
            Logging("=> 2. ADD  File - wwhisper successfully")
        else:
            ValidateFailResultAndSystem("=> Crash App")
            exit(0)
    except WebDriverException:
        ValidateFailResultAndSystem("=> Crash App")
        exit(0)
    
    icon_send_wwhisper = WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.XPATH, data["hanbiro_talk_ios"]["send_whisper"])))
    icon_send_wwhisper.click()
    Logging("=> Save Whissper successfully")
    time.sleep(5)

    try:
        select_tab_whisper=driver.find_element_by_ios_class_chain(data["hanbiro_talk_ios"]["click_tab_whisper"])
        if  select_tab_whisper.is_displayed():
            Logging("=> No Crash App") 
            TestCase_LogResult(**data["testcase_result"]["hanbiro_talk_ios"]["send_whisper"]["pass"])
            Logging("=> 3. Send Whisper successfully") 
        else:
            ValidateFailResultAndSystem("=> Crash App")
            exit(0)
    except WebDriverException:
        ValidateFailResultAndSystem("=> Send Whisper Fail")
        TestCase_LogResult(**data["testcase_result"]["hanbiro_talk_ios"]["send_whisper"]["fail"])
        exit(0)

    time.sleep(1)



    Logging("------------------------------------------------------Whisper - View Whisper -----------------------------------------------------")
    try:
        check_select_whisper = WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, data["hanbiro_talk_ios"]["select_whisper_view"])))
        check_select_whisper.click()
        Logging("=> 1. Select Whisper successfully") 
        time.sleep(6)
        content_whisper_send = driver.find_element_by_ios_class_chain(data["hanbiro_talk_ios"]["check_content_sent_whiper"])
        if  content_whisper_send.is_displayed():
            Logging("=> No Crash App") 
            Logging("=> 1. Show data Whisper successfully") 
            TestCase_LogResult(**data["testcase_result"]["hanbiro_talk_ios"]["view_whisper"]["pass"])
         
        else:
            ValidateFailResultAndSystem("=> Crash App")
            TestCase_LogResult(**data["testcase_result"]["hanbiro_talk_ios"]["view_whisper"]["fail"])
            exit(0)
    except WebDriverException:
        ValidateFailResultAndSystem("=> Crash App")
        exit(0)
    time.sleep(1)


    
    Logging("------------------------------------------------------Whisper - Download File -----------------------------------------------------")

    par_whisper = driver.find_element_by_ios_predicate('type == "XCUIElementTypeTable"')
    Logging("=> OK OK OK ")
    time.sleep(5)
    img_nu_wp= par_whisper.find_elements_by_xpath("//XCUIElementTypeCell")
    print("=> OK OK OK  nmnm",len(img_nu_wp))
    i = len(img_nu_wp)
    time.sleep(3)
    while i > 1 :
        time.sleep(5)
        h= "//XCUIElementTypeApplication[@name='HanbiroTalk']/XCUIElementTypeWindow[1]/XCUIElementTypeOther/XCUIElementTypeOther/XCUIElementTypeOther/XCUIElementTypeOther/XCUIElementTypeOther/XCUIElementTypeOther/XCUIElementTypeOther/XCUIElementTypeOther/XCUIElementTypeOther/XCUIElementTypeOther/XCUIElementTypeOther/XCUIElementTypeTable/XCUIElementTypeCell["

        h1 = str(i) + "]/XCUIElementTypeOther"
        y= h+h1
        driver.find_element_by_xpath(y).click()    
        time.sleep(1)
        print("i",i)
        i = i - 1
        break

    click_download_image_in_wwhisper=driver.find_element_by_ios_class_chain(data["hanbiro_talk_ios"]["click_image_download_whisper"])
    click_download_image_in_wwhisper.click()
    time.sleep(1)
    Logging("=> Click icon Download file successfully")



    '''
    try:
        view_image_in_message = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, data["hanbiro_talk_ios"]["click_icon_download_image_in_message"])))
        if  view_image_in_message.is_displayed():
            Logging("=> No Crash App") 
            TestCase_LogResult(**data["testcase_result"]["hanbiro_talk_ios"]["download_file_message"]["pass"])
            #Logging("=> View File Image successfully")
        else:
            ValidateFailResultAndSystem("=> Crash App")
            TestCase_LogResult(**data["testcase_result"]["hanbiro_talk_ios"]["download_file_message"]["fail"])
            exit(0)
    except WebDriverException:
        ValidateFailResultAndSystem("=> Crash App")
        exit(0)
    '''
    

    click_done_view_image=driver.find_element_by_ios_class_chain(data["hanbiro_talk_ios"]["click_done_image_download_whisper"])
    click_done_view_image.click()
    Logging("=> Click icon Done download file successfully")
    time.sleep(2)
    
    '''
    click_done_view_image=driver.find_element_by_ios_class_chain(data["hanbiro_talk_ios"]["click_done_image_download_whisper"])
    if  click_done_view_image.is_displayed():
        Logging("=> No Crash App") 
        click_done_view_image=driver.find_element_by_ios_class_chain(data["hanbiro_talk_ios"]["click_done_image_download_whisper"])
        click_done_view_image.click()
        Logging("=> Click icon Done download fail")
        TestCase_LogResult(**data["testcase_result"]["hanbiro_talk_ios"]["download_file_whisper"]["pass"])
    else:
        ValidateFailResultAndSystem("=> Crash App")
        exit(0)
    '''
  



    time.sleep(1)
    Logging("------------------------------------------------------Whisper - Forward Whisper -----------------------------------------------------")
    try:
        click_icon_forward=driver.find_element_by_ios_class_chain(data["hanbiro_talk_ios"]["icon_forward_whisper"])
        click_icon_forward.click()
        time.sleep(4)
        enter_content_write_wwhisper=driver.find_element_by_ios_class_chain(data["hanbiro_talk_ios"]["txt_input_content_whisper"])
        enter_content_write_wwhisper.click()
        enter_content_write_wwhisper.send_keys(data["hanbiro_talk_ios"]["content_forward_whisper"]+"\n")
        Logging("Input content  => pass")
        time.sleep(2)
        Logging("1. Input Content Whisper  successfully")
        icon_add_recipient = driver.find_element_by_ios_class_chain(data["hanbiro_talk_ios"]["write_whisper_add_contact"])
        if  icon_add_recipient.is_displayed():
            Logging("=> No Crash App")        
            icon_add_recipient= driver.find_element_by_ios_class_chain(data["hanbiro_talk_ios"]["write_whisper_add_contact"])
            icon_add_recipient.click()
            time.sleep(1)
            txt_search_contact_wwhisper= WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["hanbiro_talk_ios"]["txt_search_contact"])))
            txt_search_contact_wwhisper.click()
            time.sleep(2)
            Logging("BBBB")
            input_search_contact_org=driver.find_element_by_ios_class_chain(data["hanbiro_talk_ios"]["txt_input_search_contact"])
            input_search_contact_org.click()
            input_search_contact_org.send_keys(data["hanbiro_talk_ios"]["search_contact_org_talk"]+"\n")
            Logging("Input user => pass")
            time.sleep(1)
            select_contact_org=driver.find_element_by_ios_class_chain(data["hanbiro_talk_ios"]["select_contact_org"])
            select_contact_org.click()
            time.sleep(1)
            icon_confirm_contact_recippient_whisper = WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.XPATH, data["hanbiro_talk_ios"]["icon_confirm_contact_recipient_whisper"])))
            icon_confirm_contact_recippient_whisper.click()
            Logging("=> 5. Click Icon Confirm Contact successfully") 
            time.sleep(1)
            icon_send_wwhisper = WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.XPATH, data["hanbiro_talk_ios"]["send_whisper"])))
            icon_send_wwhisper.click()
            Logging("=> Save Whissper successfully")
            time.sleep(3)
        else:
                ValidateFailResultAndSystem("=> Crash App")
                exit(0)
    except WebDriverException:
        ValidateFailResultAndSystem("=> Crash App")
        TestCase_LogResult(**data["testcase_result"]["hanbiro_talk_ios"]["forward_whisper"]["fail"])
        exit(0)  
    time.sleep(1)
    driver.back()
    time.sleep(1)
    

    Logging("------------------------------------------------------Whisper - Search Whisper -----------------------------------------------------")
    try:
        txt_search_content=driver.find_element_by_ios_class_chain(data["hanbiro_talk_ios"]["txt_search_content_whisper"])
        txt_search_content.click()
        Logging("Input content search whisper => pass")
        time.sleep(2)
        Logging("1. Input Content Whisper  successfully")
        input_content_search_whisper=driver.find_element_by_ios_class_chain(data["hanbiro_talk_ios"]["input_content_whisper"])
        input_content_search_whisper.click()
        input_content_search_whisper.send_keys(data["hanbiro_talk_ios"]["content_search_whisper"]+"\n")
        Logging("Input content search whisper => pass")
        time.sleep(2)
        check_data_search_wwhisper=driver.find_element_by_ios_class_chain(data["hanbiro_talk_ios"]["check_data_search_content_wwhisper"])
        if  check_data_search_wwhisper.is_displayed():
            Logging("=> QQQQ Search successfully") 
            TestCase_LogResult(**data["testcase_result"]["hanbiro_talk_ios"]["search_whisper"]["pass"])
        else:
            ValidateFailResultAndSystem("=> Search Fail")
            TestCase_LogResult(**data["testcase_result"]["hanbiro_talk_ios"]["search_whisper"]["fail"])
            exit(0)
        time.sleep(1)
        time.sleep(1)
        driver.back()
        time.sleep(1)
        
    except WebDriverException:
        Logging("1. Search Whisper  successfully")
        exit(0)
    time.sleep(2)
    
    

    Logging("------------------------------------------------------Whisper - Filter Whisper -----------------------------------------------------")
    try:
        #icon_filter_wwhisper=driver.find_element_by_ios_class_chain(data["hanbiro_talk_ios"]["click_icon_filter_whisper"])
        #icon_filter_wwhisper.click()
        #time.sleep(2)
        Logging("click icon filter => pass")
        txt_search_content=driver.find_element_by_ios_class_chain(data["hanbiro_talk_ios"]["txt_search_content_whisper"])
        txt_search_content.click()
        Logging("Input content search whisper => pass")
        time.sleep(2)
        Logging("1. Input Content Whisper  successfully")
        input_content_search_whisper=driver.find_element_by_ios_class_chain(data["hanbiro_talk_ios"]["input_content_whisper"])
        input_content_search_whisper.click()
        input_content_search_whisper.send_keys(data["hanbiro_talk_ios"]["content_search_whisper"])
        Logging("Input content search whisper => pass")
        time.sleep(2)
        select_filter_content=driver.find_element_by_ios_class_chain(data["hanbiro_talk_ios"]["value_content_filter_whisper"])
        select_filter_content.click()
        time.sleep(3)
        btn_apply_search_whisper=driver.find_element_by_ios_class_chain(data["hanbiro_talk_ios"]["btn_apply_search_whisper"])
        btn_apply_search_whisper.click()
        time.sleep(3)
        check_data_search_wwhisper=driver.find_element_by_ios_class_chain(data["hanbiro_talk_ios"]["check_data_search_content_wwhisper"])
        if  check_data_search_wwhisper.is_displayed():
            Logging("=> Filter Whisper successfully") 
            TestCase_LogResult(**data["testcase_result"]["hanbiro_talk_ios"]["filter_whisper"]["pass"])
        else:
            ValidateFailResultAndSystem("=> Search Fail")
            TestCase_LogResult(**data["testcase_result"]["hanbiro_talk_ios"]["filter_whisper"]["fail"])
            exit(0)



        '''
        show_button_search_all = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, data["hanbiro_talk_ios"]["btn_all_filter_whisper"])))
        if  show_button_search_all.is_displayed():
            Logging("=> No Crash App") 
            click_icon_confirm_filter_whisper = WebDriverWait(driver, 50).until(EC.presence_of_element_located((By.ID, data["hanbiro_talk_ios"]["icon_confirm_filter_whisper"])))
            click_icon_confirm_filter_whisper.click()
            Logging("=> Filter Whisper successfully") 
            TestCase_LogResult(**data["testcase_result"]["hanbiro_talk_ios"]["filter_whisper"]["pass"])
        else:
            ValidateFailResultAndSystem("=> Crash App")
            TestCase_LogResult(**data["testcase_result"]["hanbiro_talk_ios"]["filter_whisper"]["fail"])
            exit(0)
        '''
    except WebDriverException:
        ValidateFailResultAndSystem("=> Filter Whisper Fail")
        exit(0)
    time.sleep(1)
    driver.back()
    time.sleep(1)

def hanbiro_talk_ios_organization():
    #hanbirotalk = CheckPresenceOfAdminsubmenu()

    try:
        hanbiro_talk_ios_organization_search()
        Logging("=> Search organization successfully")
    except WebDriverException:
        Logging("=> Search organization Fail ")
        
    try:
        hanbiro_talk_ios_message_input_msg()
        Logging("=> Input Content successfully")
    except WebDriverException:
        Logging("=> Input Content Fail ")
    hanbiro_talk_ios_message_attach_file()
    hanbiro_talk_android_message_fw_copy()

    '''   
    try:
        hanbiro_talk_ios_message_attach_file()
        Logging("=> Attach File successfully")
    except WebDriverException:
        Logging("=> Attach File Fail ")
    
    try:
        hanbiro_talk_android_message_fw_copy()
        Logging("=> Foward-Copy successfully")
    except WebDriverException:
        Logging("=> Foward-Copy Fail")
    '''

def hanbiro_talk_ios_board():
    
    hanbiro_talk_ios_write_board()

    '''
    try:
        hanbiro_talk_ios_write_board()
        Logging("=> Write Board successfully")
    except WebDriverException:
        Logging("=> Write Board Fail ")
    '''
    hanbiro_talk_ios_board_view_download_file()

    '''
    try:
        hanbiro_talk_ios_board_view_download_file()
        Logging("=> View/Download File in  Board successfully")
    except WebDriverException:
        Logging("=> View/Download File in  Board Fail ")
    '''    

    try:
        hanbiro_talk_ios_edit_board()
        Logging("=> Edit Board successfully")
    except WebDriverException:
        Logging("=> Edit Board Fail ")
    
    try:
        hanbiro_talk_ios_delete_board()
        Logging("=> Delete Board successfully")
    except WebDriverException:
        Logging("=> Delete Board Fail ")


def whisper_hanbiro_talk_ios():
    
    try:
        whisper_hanbiro_talk_ios_send_whisper()
        Logging("=> Send Whisper successfully")
    except WebDriverException:
        Logging("=> Send Whisper Fail ")
        
    








    
    











#log_in_mail_app()
#send_mail_app()
#vacation_auto_replies_mail_app()
#auto_sort_mail_app()
